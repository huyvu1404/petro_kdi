from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from io import BytesIO
from src.kdi.constants import CHANNEL_MAP
from src.utils import sanitize_excel_values
import pandas as pd

def create_labels_column(row):
    if row["Labels1"]:
        return row["Labels1"]
    elif row["Labels2"]:
        return row["Labels2"]
    elif row["Labels3"]:
        return row["Labels3"]
    elif row["Labels4"]:
        return row["Labels4"]
    else:
        return ""

def export_to_excel(df):
    try:
        # Sanitize data before processing
        df = sanitize_excel_values(df)
        
        df['PublishedDate'] = pd.to_datetime(df['PublishedDate']).dt.normalize()
        df.fillna(" ")
        df["Labels"] = df.apply(create_labels_column, axis=1)
        df = df.sort_values(by=["Sentiment", "PublishedDate"], ascending=[False, True])
        topic = df['Topic'].unique()

        selected_columns = [
            "Title", "Content", "Description", "UrlComment", "PublishedDate",
            "Sentiment", "SiteName", "Channel", "Author", "UrlTopic", "Labels", "Type",
            "Channel Group"
        ]
        for col in selected_columns:
            if col not in df.columns:
                df[col] = " "

        vietnamese_columns = {
            "Title": "Bài đăng",
            "Content": "Nội dung",
            "Description": "Mô tả",
            "UrlComment": "Link bình luận",
            "PublishedDate": "Ngày",
            "Sentiment": "Sắc thái",
            "SiteName": "Nguồn đăng",
            "Channel": "Kênh",
            "Author": "Tác giả",
            "UrlTopic": "Link bài đăng",
            "Labels": "Từ khoá",
            "Type": "Loại",
            "Channel Group": "Kênh Mới"
        }

        df = df[selected_columns].rename(columns=vietnamese_columns)
        channels = df["Kênh Mới"].unique()

        wb = Workbook()
        wb.remove(wb.active)

        for channel in channels:
            ws = wb.create_sheet(title=str(CHANNEL_MAP.get(channel, channel)))
            channel_df = df[df["Kênh Mới"] == channel].copy()

            columns = channel_df.columns.tolist()
            columns.remove("Kênh Mới")

            channel_df["STT"] = range(1, len(channel_df) + 1)
            channel_df = channel_df[["STT"] + columns]

            ws.append(channel_df.columns.tolist())
            for r in channel_df.itertuples(index=False, name=None):
                # Replace None/NaN with space to prevent cell overflow
                row_data = [" " if pd.isna(val) or val == "" else val for val in r]
                ws.append(row_data)

            header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            header_font = Font(bold=True, color="000000")
            header_align = Alignment(horizontal="center", vertical="center")

            positive_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  
            negative_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid") 

            data_align = Alignment(vertical="top", wrap_text=False)
            thin = Side(border_style="thin", color="000000")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            
            max_row = ws.max_row
            max_col = ws.max_column
            # ===== Format date column =====
            date_col_idx = None
            for idx, col in enumerate(ws[1], start=1):
                if col.value == "Ngày":
                    date_col_idx = idx
                    break

            if date_col_idx:
                for row in ws.iter_rows(
                    min_row=2,
                    max_row=max_row,
                    min_col=date_col_idx,
                    max_col=date_col_idx
                ):
                    for cell in row:
                        cell.number_format = "DD/MM/YYYY"

            row_height = 15
            column_widths = [10, 25, 22, 25, 18, 17, 10, 20, 10, 20, 18, 25, 20]

            max_row = ws.max_row
            max_col = ws.max_column

            for cell in ws[1]:
                cell.alignment = header_align
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border

            sentiment_col_idx = None
            for idx, col in enumerate(ws[1], start=1):
                if col.value == "Sắc thái":
                    sentiment_col_idx = idx
                    break

            # Center align specific columns
            center_columns = {"STT", "Ngày", "Sắc thái", "Loại", "Kênh"}
            center_col_indices = []
            for idx, col in enumerate(ws[1], start=1):
                if col.value in center_columns:
                    center_col_indices.append(idx)

            for row in ws.iter_rows(min_row=2, max_row=max_row, max_col=max_col):
                for idx, cell in enumerate(row, start=1):
                    cell.border = border
                    
                    # Center align specific columns
                    if idx in center_col_indices:
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                    else:
                        cell.alignment = data_align

                if sentiment_col_idx:
                    sentiment_value = row[sentiment_col_idx - 1].value
                    if sentiment_value == "Positive":
                        for cell in row:
                            cell.fill = positive_fill
                    elif sentiment_value == "Negative":
                        for cell in row:
                            cell.fill = negative_fill
                    
            # Remove the old STT center alignment code since it's now handled above

            for i, col_width in enumerate(column_widths, start=1):
                col_letter = ws.cell(row=1, column=i).column_letter
                ws.column_dimensions[col_letter].width = col_width

            for row_idx in range(1, max_row + 1):
                ws.row_dimensions[row_idx].height = row_height

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return buffer, topic
    except Exception as e:
        print(e)
