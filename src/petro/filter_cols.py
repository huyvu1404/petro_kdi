from src.utils import unzip_file, zip_file, sanitize_excel_values
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from fastapi import HTTPException
from datetime import datetime, timedelta

def process_zip_excel(
    zip_path: str,
    work_dir: str,
    selected_columns: list[str],
) -> str:

    # ---------- Unzip ----------
    unzip_file(zip_path, work_dir)

    # ---------- Find xlsx ----------
    excel_file = None
    for root, _, files in os.walk(work_dir):
        for name in files:
            if name.endswith(".xlsx"):
                excel_file = os.path.join(root, name)
                break
        if excel_file:
            break

    if not excel_file:
        raise HTTPException(
            status_code=400,
            detail="Không tìm thấy file xlsx trong zip"
        )

    # ---------- Read Excel ----------
    df = pd.read_excel(excel_file)
    df["PublishedDate"] = pd.to_datetime(
        df["PublishedDate"],
        errors="coerce"
    )
    df = df[
        (df["Sentiment"].str.strip() == "Neutral") &
        (df["Channel"].str.strip() == "News")
    ]

    # ---------- Validate columns ----------
    missing_cols = set(selected_columns) - set(df.columns)
    if missing_cols:
        raise HTTPException(
            status_code=400,
            detail=f"Thiếu cột: {', '.join(missing_cols)}"
        )

    df_new = df[selected_columns]
    df_new = df_new.sort_values(by="PublishedDate", ascending=True)
    
    # Add empty column at the end to prevent overflow
    df_new[" "] = " "


    def format_run_at(dt, hour, start=True):
        d = dt.replace(hour=hour, minute=0, second=0, microsecond=0)
        time_label = f"{hour}AM" if hour < 12 else f"{hour}PM"
        if start:
            return d.strftime(f"{time_label} %d.%m")
        else:
            return d.strftime(f"{time_label} %d.%m.%Y")

    def get_run_at(now=None):
        if now is None:
            now = datetime.now()

        if now.hour >= 11:
            
            start = format_run_at(now, 11)
            end = format_run_at(now, 16, False)
        else:

            yesterday = now - timedelta(days=1)
            start = format_run_at(yesterday, 16)
            end = format_run_at(now, 11, False)

        return f"{start} to {end}"

    output_xlsx = os.path.join(work_dir, f"Ngành xăng dầu_News Tổng hợp tin trung lập {get_run_at()}.xlsx")
    df_new = sanitize_excel_values(df_new)
    # Fill empty cells with space to prevent overflow
    df_new = df_new.fillna(" ")
    df_new = df_new.replace("", " ")
    df_new.to_excel(output_xlsx, index=False)

    # ---------- Style + Auto width ----------
    wb = load_workbook(output_xlsx)
    ws = wb.active
    thin_black_border = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )
    header_map = {
        cell.value: cell.column
        for cell in ws[1]
    }

    header_font = Font(bold=True)
    header_fill = PatternFill(
        start_color="FFFF00",
        end_color="FFFF00",
        fill_type="solid",
    )

    # Header style (exclude last column - empty column)
    for idx, cell in enumerate(ws[1], start=1):
        if idx < ws.max_column:  # Skip last column
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Data style
    center_columns = {"Topic", "Sentiment", "PublishedDate"}
    for col_name in center_columns:
        if col_name in header_map:
            col_letter = get_column_letter(header_map[col_name])
            for cell in ws[col_letter][1:]:
                cell.alignment = Alignment(horizontal="center", vertical="center")
    # Date format
    if "PublishedDate" in header_map:
        col_letter = get_column_letter(header_map["PublishedDate"])
        for cell in ws[col_letter][1:]:
            cell.number_format = "DD/MM/YYYY"            
    # Auto column width (exclude last column)
    for col_idx, col in enumerate(ws.iter_cols(), start=1):
        if col_idx < ws.max_column:  # Skip last column
            max_length = 0
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].width = 20
    
    max_row = ws.max_row
    max_col = ws.max_column

    # Apply border (exclude last column)
    for row in ws.iter_rows(min_row=1, max_row=max_row,
                            min_col=1, max_col=max_col - 1):
        for cell in row:
            cell.border = thin_black_border

    wb.save(output_xlsx)

    # ---------- Zip result ----------
    output_zip = os.path.join(work_dir, "result.zip")
    zip_file(output_xlsx, output_zip)

    return output_zip

